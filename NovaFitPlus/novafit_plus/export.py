import os, json, datetime as dt
from .db import get_conn

def export_json(db_path: str, export_dir: str, user_name: str, days: int = 14) -> str:
    os.makedirs(export_dir, exist_ok=True)
    start = (dt.date.today() - dt.timedelta(days=days)).isoformat()
    data = {"profile": {}, "activities": [], "weather": [], "water_intake": []}

    with get_conn(db_path) as c:
        cur = c.cursor()
        cur.execute("SELECT id, name, age, sex, height_cm, weight_kg, activity_level FROM users WHERE name=?", (user_name,))
        u = cur.fetchone()
        if u:
            uid, name, age, sex, h, w, al = u
            data["profile"] = {"name": name, "age": age, "sex": sex, "height_cm": h, "weight_kg": w, "activity_level": al}
        else:
            uid = None

        if uid is not None:
            cur.execute('''SELECT date, steps, calories, mood, notes, sleep_hours FROM activities
                           WHERE user_id=? AND date >= ? ORDER BY date ASC''', (uid, start))
            for d, st, cal, mood, notes, sl in cur.fetchall():
                data["activities"].append({"date": d, "steps": st, "calories": cal, "mood": mood, "notes": notes, "sleep_hours": sl})

            cur.execute('''SELECT date, ml, source, created_at FROM water_intake
                           WHERE user_id=? AND date >= ? ORDER BY id ASC''', (uid, start))
            for d, ml, src, ts in cur.fetchall():
                data["water_intake"].append({"date": d, "ml": ml, "source": src, "timestamp": ts})

        cur.execute('''SELECT date, city, latitude, longitude, temp_max, temp_min, humidity, wind_speed, condition
                       FROM weather WHERE date >= ? ORDER BY date ASC''', (start,))
        for d, city, lat, lon, tmax, tmin, hum, wind, cond in cur.fetchall():
            data["weather"].append({
                "date": d, "city": city, "lat": lat, "lon": lon,
                "temp_max": tmax, "temp_min": tmin, "humidity": hum, "wind_speed": wind, "condition": cond
            })

    out_path = os.path.join(export_dir, "novafit_plus_report.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return out_path

def export_excel(db_path: str, export_dir: str, user_name: str, days: int = 14) -> str:
    from openpyxl import Workbook
    os.makedirs(export_dir, exist_ok=True)
    start = (dt.date.today() - dt.timedelta(days=days)).isoformat()

    wb = Workbook()
    ws_profile = wb.active; ws_profile.title = "profile"
    ws_acts = wb.create_sheet("activities")
    ws_weather = wb.create_sheet("weather")
    ws_water = wb.create_sheet("water_intake")
    ws_hyd = wb.create_sheet("hydration_summary")
    ws_sleep = wb.create_sheet("sleep_summary")

    with get_conn(db_path) as c:
        cur = c.cursor()
        # Profile
        cur.execute("SELECT id, name, age, sex, height_cm, weight_kg, activity_level FROM users WHERE name=?", (user_name,))
        u = cur.fetchone()
        ws_profile.append(["field", "value"])
        if u:
            uid, name, age, sex, h, w, al = u
            for k, v in [("name", name), ("age", age), ("sex", sex), ("height_cm", h), ("weight_kg", w), ("activity_level", al)]:
                ws_profile.append([k, v])
        else:
            uid = None
            h, w = 170, 70  # defaults for summaries

        # Activities
        ws_acts.append(["date", "steps", "calories", "mood", "notes", "sleep_hours"])
        if uid is not None:
            cur.execute('''SELECT date, steps, calories, mood, notes, sleep_hours FROM activities
                           WHERE user_id=? AND date >= ? ORDER BY date ASC''', (uid, start))
            for row in cur.fetchall():
                ws_acts.append(list(row))

        # Weather
        ws_weather.append(["date", "city", "lat", "lon", "temp_max", "temp_min", "humidity", "wind_speed", "condition"])
        cur.execute('''SELECT date, city, latitude, longitude, temp_max, temp_min, humidity, wind_speed, condition
                       FROM weather WHERE date >= ? ORDER BY date ASC''', (start,))
        for row in cur.fetchall():
            ws_weather.append(list(row))

        # Water intake
        ws_water.append(["date", "ml", "source", "timestamp"])
        if uid is not None:
            cur.execute('''SELECT date, ml, source, created_at FROM water_intake
                           WHERE user_id=? AND date >= ? ORDER BY id ASC''', (uid, start))
            for row in cur.fetchall():
                ws_water.append(list(row))

        # Hydration summary per day
        from .hydration import daily_water_goal_ml
        from .db import daily_water_total
        start_date = dt.date.today() - dt.timedelta(days=days-1)
        ws_hyd.append(["date", "goal_ml", "total_ml", "pct", "tmax"])
        for i in range(days):
            d = (start_date + dt.timedelta(days=i)).isoformat()
            goal = daily_water_goal_ml(w, d, db_path, user_name)
            total = daily_water_total(db_path, user_name, d)
            pct = 0 if goal <= 0 else round(total * 100 / goal, 1)
            cur.execute("SELECT temp_max FROM weather WHERE date=? ORDER BY id DESC LIMIT 1", (d,))
            wr = cur.fetchone()
            tmax = wr[0] if wr else None
            ws_hyd.append([d, goal, total, pct, tmax])

        # Sleep summary per day
        ws_sleep.append(["date", "sleep_hours", "percent_vs_8h"])
        if uid is not None:
            cur.execute('''SELECT date, sleep_hours FROM activities
                           WHERE user_id=? AND date >= ? ORDER BY date ASC''', (uid, start))
            for d, sl in cur.fetchall():
                pct = round((float(sl or 0) / 8.0) * 100, 1) if sl else 0
                ws_sleep.append([d, sl, pct])

    out_path = os.path.join(export_dir, "novafit_plus_report.xlsx")
    wb.save(out_path)
    return out_path


def export_csv(db_path: str, export_dir: str, user_name: str, days: int = 14) -> dict:
    import csv, datetime as dt
    os.makedirs(export_dir, exist_ok=True)
    start = (dt.date.today() - dt.timedelta(days=days)).isoformat()
    outs = {}
    with get_conn(db_path) as c:
        cur = c.cursor()
        cur.execute("SELECT id FROM users WHERE name=?", (user_name,))
        r = cur.fetchone()
        uid = r[0] if r else None

        # activities.csv
        act_path = os.path.join(export_dir, "activities.csv"); outs["activities"] = act_path
        with open(act_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["date","steps","calories","mood","notes","sleep_hours"])
            if uid is not None:
                cur.execute('''SELECT date, steps, calories, mood, notes, sleep_hours FROM activities
                               WHERE user_id=? AND date >= ? ORDER BY date ASC''', (uid, start))
                for row in cur.fetchall(): w.writerow(row)

        # water_intake.csv
        w_path = os.path.join(export_dir, "water_intake.csv"); outs["water_intake"] = w_path
        with open(w_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["date","ml","source","timestamp"])
            if uid is not None:
                cur.execute('''SELECT date, ml, source, created_at FROM water_intake
                               WHERE user_id=? AND date >= ? ORDER BY id ASC''', (uid, start))
                for row in cur.fetchall(): w.writerow(row)

        # weather.csv
        wx_path = os.path.join(export_dir, "weather.csv"); outs["weather"] = wx_path
        with open(wx_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["date","city","lat","lon","temp_max","temp_min","humidity","wind_speed","condition"])
            cur.execute('''SELECT date, city, latitude, longitude, temp_max, temp_min, humidity, wind_speed, condition
                           FROM weather WHERE date >= ? ORDER BY date ASC''', (start,))
            for row in cur.fetchall(): w.writerow(row)

    return outs
